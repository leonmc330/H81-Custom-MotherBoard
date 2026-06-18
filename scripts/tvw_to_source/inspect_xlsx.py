"""
Inspect .xlsx structure for designing the rules converter.

Usage:
    python inspect_xlsx.py <file.xlsx>                # survey: sheet list
    python inspect_xlsx.py <file.xlsx> --sheet NAME   # full dump of one sheet
    python inspect_xlsx.py <file.xlsx> --all          # dump every sheet
"""
import argparse
from pathlib import Path
import openpyxl


def _color_hex(cell):
    try:
        rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
        if rgb and rgb not in ("00000000", "FFFFFFFF", None) and not str(rgb).startswith("00"):
            return f"[{str(rgb)[-6:]}]"
    except Exception:
        pass
    return ""


def survey(wb, name):
    print(f"\n=== {name} ===")
    print(f"Sheets ({len(wb.worksheets)}):")
    for s in wb.worksheets:
        n_comments = sum(1 for row in s.iter_rows() for c in row if c.comment)
        print(
            f"  {s.title!r:40} rows={s.max_row:4} cols={s.max_column:3} "
            f"merged={len(s.merged_cells.ranges):4} comments={n_comments}"
        )


def dump_sheet(wb, sheet_name):
    sheet = wb[sheet_name]
    print(f"\n--- sheet: {sheet_name!r} (dim {sheet.dimensions}) ---")

    print(f"\nmerged ranges ({len(sheet.merged_cells.ranges)}):")
    for r in sorted(sheet.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
        v = sheet.cell(r.min_row, r.min_col).value
        s = (str(v) if v is not None else "").replace("\n", " ").strip()
        print(f"  {r}: {s[:80]!r}")

    print(f"\nrows:")
    for r in range(1, sheet.max_row + 1):
        parts = []
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(r, c)
            v = cell.value
            if v is None:
                txt = "-"
            else:
                s = str(v).strip().replace("\n", " / ")
                txt = (s[:60] + "...") if len(s) > 60 else s
            parts.append(txt + _color_hex(cell))
        print(f"R{r:3}: " + " | ".join(parts))

    comments = [
        (c.coordinate, c.comment.text)
        for row in sheet.iter_rows()
        for c in row
        if c.comment
    ]
    if comments:
        print(f"\ncomments ({len(comments)}):")
        for coord, text in comments:
            print(f"  {coord}: {text[:200].replace(chr(10), ' / ')!r}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    ap.add_argument("file")
    ap.add_argument("--sheet", default=None, help="Dump one sheet by name")
    ap.add_argument("--all", action="store_true", help="Dump every sheet")
    args = ap.parse_args()

    path = Path(args.file)
    wb = openpyxl.load_workbook(path, data_only=True)

    if args.sheet:
        dump_sheet(wb, args.sheet)
    elif args.all:
        survey(wb, path.name)
        for s in wb.worksheets:
            dump_sheet(wb, s.title)
    else:
        survey(wb, path.name)


if __name__ == "__main__":
    main()
