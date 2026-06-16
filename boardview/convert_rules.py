# SPDX-License-Identifier: LGPL-3.0-or-later
# Copyright (C) 2026 Thermetery Technology LLC

"""
Translate power-sequencing diagnostic spreadsheets to YAML.

Runs entirely locally. The .xlsx contents never leave this machine.

Each sheet within a workbook is treated as one motherboard generation/platform
(the sheet name becomes the platform key). Multiple input files are merged
into a single output keyed by platform.

Usage:
    python convert_rules.py <input.xlsx> [<input2.xlsx> ...] -o <output.yaml>

Example:
    python convert_rules.py private/sequence_a.xlsx private/sequence_b.xlsx \\
        -o private/rules.yaml
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import yaml


# Cell fill colors carry meaning in the source spreadsheet.
COLOR_SEMANTICS = {
    "FFFF0000": "critical_rail",      # red
    "FFFFFF00": "control_signal",     # yellow
    "FF00FF00": "target_rail",        # bright green
    "FF92D050": "target_rail",        # alt green
    "FF00B050": "target_rail",        # darker green
    "FF00B0F0": "metadata_highlight", # cyan (used for "data source" notes)
}

# Header cells anchor column layout (and are skipped wherever they repeat).
# Maps header text → field name. Multiple labels can map to the same field
# because Mr. Ren uses slight variations across sheets.
HEADER_LABELS = {
    "上电时序": "stage",
    "内存不过检修流程": "stage",
    "电压或信号标准名称": "signal",
    "电压或信号标准名称/固定测试点": "signal",
    "正常工作电压": "voltage",
    "正常对地值参考": "resistance",
}

# Section terminator phrases. The matched row's text becomes the section's
# diagnosis_summary; the id is for stable references in downstream tooling.
# Order matters: more specific patterns first.
SECTION_TERMINATOR_PATTERNS = [
    ("no_trigger",
     "以上电压或信号出故障会引起不触发"),
    ("post_trigger_fault",
     "以上电压或信号出现故障会引起掉电、缺电压、缺时钟、缺复位"),
    ("memory_not_detected",
     "以上内存工作条件出现故障会引不认内存"),
    ("memory_bus_fault",
     "以上内存总线故障会引起内存检测不通过"),
    ("pcie_x16_no_display",
     "以上电压或信号不正常会引起PCIE_X16不显示"),
    ("cpu_not_addressing_done",
     "CPU不寻址按以上步骤操作"),
]

# Private/contact-info markers — rows matching these are dropped entirely.
# Mr. Ren obfuscates ("++加++微++信" instead of "加微信") so we match loosely.
PRIVATE_CONTACT_PATTERNS = [
    re.compile(r"微+\W*信"),
    re.compile(r"\bQQ\b", re.IGNORECASE),
    re.compile(r"咨\s*\W*\s*询"),
    re.compile(r"教\s*\W*\s*学"),
    re.compile(r"课\s*\W*\s*程"),
    re.compile(r"任\s*\W*\s*师"),
    re.compile(r"维\s*修\s*培\s*训"),
    re.compile(r"全套系统化"),
    re.compile(r"\d{8,}"),  # phone / WeChat IDs
]


def _is_private(text):
    return any(p.search(text) for p in PRIVATE_CONTACT_PATTERNS)


def _normalize_label(text):
    """Collapse runs of whitespace (newlines, tabs, multiple spaces) into one
    space. Excel stores multi-line stage labels as 'G3状态\\n（PCH）' and we
    want them on one line."""
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def _is_likely_note(text):
    """A 'signal' cell that's actually a sentence-style note rather than a net
    name. Heuristic: contains a run of Chinese characters and is on the longer
    side. Real net names like VCCPRIM_3P3 don't carry Chinese text."""
    if not text:
        return False
    has_chinese = any("一" <= ch <= "鿿" for ch in text)
    return has_chinese and len(text) > 20


def _semantic_for_color(rgb):
    if rgb is None:
        return None
    return COLOR_SEMANTICS.get(str(rgb).upper())


def _merge_map(sheet):
    """Build {(row,col): top-left-value} so merged stage labels propagate
    to every row they cover."""
    m = {}
    for r in sheet.merged_cells.ranges:
        v = sheet.cell(r.min_row, r.min_col).value
        for row in range(r.min_row, r.max_row + 1):
            for col in range(r.min_col, r.max_col + 1):
                m[(row, col)] = v
    return m


def _is_merged_with_neighbor(sheet, row, col_a, col_b):
    """Return True if (row, col_a) and (row, col_b) are part of the same
    merged range (i.e., reading both columns yields duplicated content)."""
    for r in sheet.merged_cells.ranges:
        if (r.min_row <= row <= r.max_row
                and r.min_col <= min(col_a, col_b)
                and r.max_col >= max(col_a, col_b)):
            return True
    return False


def _cell(sheet, merges, row, col):
    raw = sheet.cell(row, col).value
    if raw is None:
        return merges.get((row, col))
    return raw


def _signal_color(cell):
    fill = getattr(cell, "fill", None)
    if not fill:
        return None
    fg = getattr(fill, "fgColor", None)
    if not fg:
        return None
    try:
        return fg.rgb
    except Exception:
        return None


def _parse_signal(text):
    """Parse the signal/net column. Preserves raw and pulls out probe location
    and voltage-group prefix when present."""
    if text is None or not str(text).strip():
        return None
    raw = str(text).strip()
    out = {"raw": raw}

    # Voltage-group prefix: "1.8V:->PCH VCCVRM/VCCDFTERM"
    m = re.match(r"^(\d+\.?\d*\s*V)\s*:\s*->\s*(.+)$", raw)
    if m:
        out["voltage_group"] = m.group(1).replace(" ", "")
        rest = m.group(2).strip()
        parts = rest.split(None, 1)
        if len(parts) == 2:
            out["component"] = parts[0]
            out["net"] = parts[1]
        else:
            out["net"] = rest
        return out

    # Probe-location form: "NET -> COMPONENT PIN[/PIN...]"
    if "->" in raw:
        net, _, probe = raw.partition("->")
        out["net"] = net.strip()
        probe = probe.strip()
        parts = probe.split(None, 1)
        if len(parts) == 2:
            comp, pins = parts
            out["probe_at"] = [
                {"component": comp.strip(), "pin": p.strip()}
                for p in pins.split("/")
            ]
        elif parts:
            out["probe_at"] = [{"component": parts[0]}]
        return out

    out["net"] = raw
    return out


def find_header_layout(sheet, merges):
    """Locate the row containing column headers and return (row, layout)
    where layout maps field names → column numbers. Looks at the first
    30 rows for resilience against varied preamble lengths."""
    for row_num in range(1, min(sheet.max_row, 30) + 1):
        layout = {}
        for col in range(1, sheet.max_column + 1):
            v = _cell(sheet, merges, row_num, col)
            if v is None:
                continue
            text = str(v).strip()
            for label, field in HEADER_LABELS.items():
                if label in text and field not in layout:
                    layout[field] = col
                    break
        if {"stage", "signal", "voltage"}.issubset(layout.keys()):
            return row_num, layout
    return None, {}


def extract_preamble(sheet, merges, header_row):
    """Pull title / source-board / URLs from rows above the header. Drops
    any row matching contact-info patterns."""
    out = {}
    for row_num in range(1, header_row):
        for col in range(1, sheet.max_column + 1):
            v = _cell(sheet, merges, row_num, col)
            if v is None:
                continue
            text = str(v).strip()
            if not text or _is_private(text):
                continue
            if "http" in text.lower():
                out.setdefault("reference_urls", []).append(text)
            elif "数据来源" in text or "板号" in text or re.search(r"MS-\d", text):
                out["source_board"] = text
            elif re.search(r"上电时序|关键信号|LGA\d|芯片组|汇总", text):
                if "title" not in out:
                    out["title"] = text
            else:
                out.setdefault("other_preamble", []).append(text)
            break  # one note per row
    return out


def find_terminator(joined_text):
    """Detect a section-terminator row. Returns (id, summary_text) or None."""
    for tid, marker in SECTION_TERMINATOR_PATTERNS:
        if marker in joined_text:
            return tid, joined_text.strip()
    # Generic fallbacks
    if re.search(r"以上.{1,40}会引起", joined_text):
        return "section_end_generic", joined_text.strip()
    if "按以上步骤操作" in joined_text:
        return "procedure_end_generic", joined_text.strip()
    return None


def extract(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheets_out = []

    for sheet in wb.worksheets:
        merges = _merge_map(sheet)
        header_row, layout = find_header_layout(sheet, merges)
        if header_row is None:
            # No tabular content (e.g., empty placeholder sheets like 'Sheet3').
            continue

        stage_col = layout["stage"]
        signal_col = layout["signal"]
        voltage_col = layout["voltage"]
        resistance_col = layout.get("resistance")  # may be absent

        preamble = extract_preamble(sheet, merges, header_row)

        sections = []
        section_idx = 1
        section = {
            "id": f"section_{section_idx}",
            "diagnosis_summary": None,
            "stages": [],
        }
        stage = None

        def _flush_stage():
            nonlocal stage
            if stage and stage["signals"]:
                section["stages"].append(stage)
            stage = None

        def _flush_section():
            nonlocal section, section_idx
            if section["stages"] or section["diagnosis_summary"]:
                sections.append(section)
            section_idx += 1
            section = {
                "id": f"section_{section_idx}",
                "diagnosis_summary": None,
                "stages": [],
            }

        for row in range(header_row + 1, sheet.max_row + 1):
            stage_val = _cell(sheet, merges, row, stage_col)
            signal_val = _cell(sheet, merges, row, signal_col)
            voltage_val = _cell(sheet, merges, row, voltage_col)
            resistance_val = (
                _cell(sheet, merges, row, resistance_col) if resistance_col else None
            )

            stage_str = _normalize_label(stage_val)
            signal_str = _normalize_label(signal_val)
            voltage_str = _normalize_label(voltage_val)
            resistance_str = _normalize_label(resistance_val)

            # Skip empty rows
            if not (stage_str or signal_str or voltage_str or resistance_str):
                continue

            # Deduplicated join — when a row is merged across all four columns,
            # _cell returns the same text in each, and we don't want N copies
            # of the terminator phrase ending up in diagnosis_summary.
            joined_unique = " ".join(
                dict.fromkeys(
                    s for s in [stage_str, signal_str, voltage_str, resistance_str] if s
                )
            )

            # Skip private/contact rows
            if _is_private(joined_unique):
                continue

            # Skip header repeats
            if any(label in stage_str or label in signal_str
                   for label in HEADER_LABELS):
                continue

            # Section terminator?
            term = find_terminator(joined_unique)
            if term:
                tid, summary = term
                section["id"] = tid
                section["diagnosis_summary"] = summary
                _flush_stage()
                _flush_section()
                continue

            # New stage?
            if stage_str and (stage is None or stage.get("label") != stage_str):
                _flush_stage()
                stage = {"label": stage_str, "signals": []}

            # If voltage and resistance came from the same merged range,
            # don't duplicate the text.
            if (voltage_str and resistance_str
                    and resistance_col
                    and voltage_str == resistance_str
                    and _is_merged_with_neighbor(sheet, row, voltage_col, resistance_col)):
                resistance_str = ""

            color_rgb = _signal_color(sheet.cell(row, signal_col))
            semantic = _semantic_for_color(color_rgb)

            # A "signal" cell that's really a Chinese-language sentence is
            # actually an inline note from Mr. Ren about the signals above.
            if signal_str and _is_likely_note(signal_str):
                if stage is None:
                    stage = {"label": "(unlabeled)", "signals": []}
                stage["signals"].append({"note": signal_str})
                continue

            # Procedural step: no signal, but text in the voltage/resistance
            # columns. Common in "CPU不寻址 检修流程" etc.
            if not signal_str and (voltage_str or resistance_str):
                step_text = voltage_str or resistance_str
                if voltage_str and resistance_str and voltage_str != resistance_str:
                    step_text = f"{voltage_str} | {resistance_str}"
                if stage is None:
                    stage = {"label": "(unlabeled)", "signals": []}
                stage["signals"].append({"step": step_text})
                continue

            signal = _parse_signal(signal_str or None)
            if signal:
                entry = {"signal": signal}
                if voltage_str:
                    entry["expected_voltage"] = voltage_str
                if resistance_str:
                    entry["resistance_to_ground"] = resistance_str
                if semantic:
                    entry["semantic"] = semantic
                if stage is None:
                    stage = {"label": "(unlabeled)", "signals": []}
                stage["signals"].append(entry)

        _flush_stage()
        _flush_section()

        sheets_out.append({
            "sheet": sheet.title,
            "metadata": preamble,
            "sections": sections,
        })

    return {"source": workbook_path.name, "sheets": sheets_out}


def main():
    ap = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    ap.add_argument("inputs", type=Path, nargs="+", help="One or more .xlsx files")
    ap.add_argument("-o", "--output", type=Path, required=True, help="Destination .yaml")
    args = ap.parse_args()

    platforms = {}
    sources = []
    for inp in args.inputs:
        if not inp.exists():
            sys.exit(f"Input not found: {inp}")
        sources.append(inp.name)
        data = extract(inp)
        for sheet in data["sheets"]:
            name = sheet["sheet"]
            if name in platforms:
                sys.exit(
                    f"Duplicate platform '{name}': appears in both "
                    f"{platforms[name]['source']} and {inp.name}. "
                    f"Rename one of the sheets, or namespace them."
                )
            platforms[name] = {
                "source": inp.name,
                "metadata": sheet["metadata"],
                "sections": sheet["sections"],
            }

    out = {"source_files": sources, "platforms": platforms}
    args.output.write_text(
        yaml.safe_dump(out, allow_unicode=True, sort_keys=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {args.output} ({len(platforms)} platform(s))")


if __name__ == "__main__":
    main()
